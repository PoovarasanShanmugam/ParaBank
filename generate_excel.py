import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_test_cases_excel():
    """
    Generates a high-quality, professional, and styled Excel sheet containing
    BDD test cases (positive and negative) for the ParaBank UI Automation suite.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases Suite"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Theme Color Palette & Fonts
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Soft Gray
    
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1F497D")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    # 1. Title Block
    ws.merge_cells("A1:H1")
    ws["A1"] = "ParaBank Account Registration & Sign-In Test Cases"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 40
    
    # Empty Spacer Row
    ws.row_dimensions[2].height = 15
    
    # 2. Headers
    headers = [
        "Test Case ID", 
        "Module", 
        "Test Case Title", 
        "Preconditions", 
        "Steps (BDD Style)", 
        "Expected Result", 
        "Test Type", 
        "Priority"
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    ws.row_dimensions[3].height = 30
    
    # 3. Test Cases Data
    test_cases = [
        (
            "TC_PB_01",
            "Registration & Login",
            "Successful User Registration, Sign-In, and Balance Retrieval",
            "ParaBank website is accessible and user is not registered.",
            "Given the user opens the ParaBank application\n"
            "When the user navigates to the registration page\n"
            "And the user registers a new account with dynamic valid credentials\n"
            "Then the registration confirmation message should contain 'Your account was created successfully. You are now logged in.'\n"
            "When the user logs out of the application\n"
            "And the user logs in with the newly registered credentials\n"
            "Then the user should see their accounts overview dashboard\n"
            "And the user logs the account balance displayed on the page",
            "1. User is registered successfully.\n"
            "2. Logout redirect succeeds.\n"
            "3. Login using newly registered credentials succeeds.\n"
            "4. Account Overview page displays, and the total balance is printed/logged to console.",
            "Positive",
            "High"
        ),
        (
            "TC_PB_02",
            "Registration",
            "Failed Registration - Password Mismatch validation",
            "Registration page is accessible.",
            "Given the user opens the ParaBank application\n"
            "When the user navigates to the registration page\n"
            "And the user attempts registration with mismatched passwords\n"
            "Then registration fails showing error 'Passwords did not match.'",
            "Registration fails on field verification and shows 'Passwords did not match.' beneath the confirm password field.",
            "Negative",
            "Medium"
        ),
        (
            "TC_PB_03",
            "Registration",
            "Failed Registration - Missing Required Fields validation",
            "Registration page is accessible.",
            "Given the user opens the ParaBank application\n"
            "When the user navigates to the registration page\n"
            "And the user attempts registration with missing first name and username fields\n"
            "Then registration fails showing error 'First name is required.' or 'Username is required.'",
            "Registration fails and proper warning messages appear beside the empty required input boxes.",
            "Negative",
            "High"
        ),
        (
            "TC_PB_04",
            "Login Panel",
            "Failed Sign-In - Invalid credentials",
            "Homepage login sidebar is visible.",
            "Given the user opens the ParaBank application\n"
            "When the user attempts login with invalid credentials 'nonexistent_test_user' and 'WrongPassword99!'\n"
            "Then login fails showing error 'The username and password could not be verified.'",
            "Sign-in fails. An error dialog banner displays stating the credentials could not be verified.",
            "Negative",
            "High"
        )
    ]
    
    # 4. Write Data & Apply Styles
    for row_idx, tc in enumerate(test_cases, 4):
        is_even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(tc, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            
            # Alternating row background for premium readability
            if is_even:
                cell.fill = gray_fill
            
            # Alignment configuration
            if col_idx in [1, 7, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                
        ws.row_dimensions[row_idx].height = 110
        
    # 5. Define Specific Column Widths
    column_widths = {
        "A": 15,  # Test Case ID
        "B": 18,  # Module
        "C": 35,  # Title
        "D": 35,  # Preconditions
        "E": 60,  # Steps (BDD)
        "F": 45,  # Expected Result
        "G": 12,  # Test Type
        "H": 12   # Priority
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    wb.save("test_cases.xlsx")
    print("[EXCEL] File 'test_cases.xlsx' successfully written.")

if __name__ == "__main__":
    generate_test_cases_excel()
