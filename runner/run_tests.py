import argparse
import os
import subprocess
import sys
import concurrent.futures

# Ensure working directory is always the project root (one level up from runner/)
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(PROJECT_ROOT)
sys.path.insert(0, PROJECT_ROOT)

class ParaBankTestRunner:
    """
    Test runner class that orchestrates:
    1. Creating the test case spreadsheet from code.
    2. Setting up reports and evidence paths.
    3. Executing Gherkin scenarios via Behave (supporting sequential & parallel modes).
    """
    def __init__(self):
        self.project_root = PROJECT_ROOT
        # Mapping of scenario tags to their category tags for parallel filtering
        self.tag_mapping = {
            "@TC_PB_01": ["@TC_PB_01", "@positive", "@registration", "@login"],
            "@TC_PB_02": ["@TC_PB_02", "@negative", "@registration"],
            "@TC_PB_03": ["@TC_PB_03", "@negative", "@registration"],
            "@TC_PB_04": ["@TC_PB_04", "@negative", "@login"]
        }
        
    def generate_excel_documentation(self):
        """
        Generates the test case spreadsheet and verifies it is populated.
        """
        print("\n[INFO] Generating test cases spreadsheet...")
        try:
            from generate_excel import generate_test_cases_excel
            generate_test_cases_excel()
            print("[INFO] test_cases.xlsx file created successfully.")
        except Exception as e:
            print(f"[ERROR] Failed to compile test cases spreadsheet: {e}")
            raise e
            
        print("[INFO] Verifying test_cases.xlsx is not empty...")
        try:
            import openpyxl
            wb = openpyxl.load_workbook("test_cases.xlsx")
            ws = wb.active
            total_rows = ws.max_row
            total_cols = ws.max_column
            
            # Row 1 is Title, Row 2 is Spacer, Row 3 is Headers. Test cases must start from Row 4.
            if total_rows <= 3:
                raise ValueError("Spreadsheet does not contain test case rows!")
                
            # Perform value assertions to make sure test cases exist
            test_case_ids = [ws.cell(row=r, column=1).value for r in range(4, total_rows + 1)]
            if not any(test_case_ids):
                raise ValueError("No valid Test Case IDs found in the spreadsheet rows!")
                
            print(f"[SUCCESS] Spreadsheet verification passed.")
            print(f"   - Sheet Name : '{ws.title}'")
            print(f"   - Total Rows : {total_rows}")
            print(f"   - Total Cols : {total_cols}")
            print(f"   - Test Cases : {', '.join(str(tc_id) for tc_id in test_case_ids if tc_id)}")
        except Exception as e:
            print(f"[ERROR] test_cases.xlsx verification failed: {e}")
            sys.exit(1)

    def prepare_directories(self):
        """
        Prepares reports and JUnit directories.
        """
        print("\n[INFO] Preparing reports directories...")
        os.makedirs("reports", exist_ok=True)
        os.makedirs("reports/junit", exist_ok=True)

    def run_tests_by_tag(self, tag: str = None) -> int:
        """
        Runs Behave scenarios filtered by the provided tag sequentially.
        """
        print("\n[INFO] Starting sequential Behave runner...")
        
        # Base behave execution flags
        cmd = [
            "behave",
            "-f", "html-pretty", "-o", "reports/report.html",
            "-f", "json", "-o", "reports/report.json",
            "-f", "pretty",
            "--junit", "--junit-directory", "reports/junit"
        ]
        
        # Apply the tag filter if provided
        if tag:
            resolved_tag = tag if tag.startswith("@") else f"@{tag}"
            cmd.extend(["--tags", resolved_tag])
            print(f"[RUNNER] Filtering by tag: '{resolved_tag}'")
        else:
            print("[RUNNER] No tag filter specified. Running all tests.")
            
        print(f"[EXEC] Command: {' '.join(cmd)}")
        
        try:
            result = subprocess.run(cmd, cwd=self.project_root, check=False)
            self._print_completion_banner(result.returncode, is_parallel=False)
            return result.returncode
        except Exception as e:
            print(f"[ERROR] Executing Behave framework encountered exception: {e}")
            raise e

    def run_tests_in_parallel(self, resolved_tag: str = None) -> int:
        """
        Discovers the scenarios matching the tag filter and executes them in parallel.
        """
        # Determine which scenario tags match the filter
        active_tags = []
        for tc_tag, mapped in self.tag_mapping.items():
            if not resolved_tag:
                active_tags.append(tc_tag)
            else:
                clean_filter = resolved_tag.replace("@", "").lower()
                if any(clean_filter in m.replace("@", "").lower() for m in mapped):
                    active_tags.append(tc_tag)
                    
        if not active_tags:
            print(f"[WARNING] No scenarios matched parallel filter: '{resolved_tag}'. Falling back to sequential execution.")
            return self.run_tests_by_tag(resolved_tag)
            
        print("\n" + "=" * 80)
        print(f"      INITIATING PARALLEL EXECUTION: {len(active_tags)} SCENARIOS RUNNING CONCURRENTLY")
        print("=" * 80)
        print(f" Target Scenarios: {', '.join(active_tags)}")
        print("-" * 80)
        
        def run_single_tag(tag):
            safe_tag = tag.replace("@", "").lower()
            # Use separate report paths to avoid file conflicts
            cmd = [
                "behave",
                "-f", "html-pretty", "-o", f"reports/report_{safe_tag}.html",
                "-f", "json", "-o", f"reports/report_{safe_tag}.json",
                "-f", "pretty",
                "--junit", "--junit-directory", f"reports/junit_{safe_tag}",
                "--tags", tag
            ]
            print(f"[LAUNCH] Thread worker started for: {tag}")
            result = subprocess.run(cmd, cwd=self.project_root, capture_output=True, text=True)
            print(f"[COMPLETE] Thread worker finished for: {tag} (Exit Code: {result.returncode})")
            return tag, result.returncode, result.stdout, result.stderr

        exit_codes = {}
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(active_tags)) as executor:
            futures = [executor.submit(run_single_tag, tag) for tag in active_tags]
            for future in concurrent.futures.as_completed(futures):
                tag, code, stdout, stderr = future.result()
                exit_codes[tag] = code
                
                print("\n" + "=" * 80)
                print(f" LOGS FOR SCENARIO WORKER: {tag}")
                print("=" * 80)
                print(stdout)
                if stderr:
                    print("--- ERROR WORKER LOGS ---")
                    print(stderr)
                print("=" * 80 + "\n")
                
        failed_tags = [t for t, c in exit_codes.items() if c != 0]
        exit_code = 1 if failed_tags else 0
        self._print_completion_banner(exit_code, is_parallel=True, details=exit_codes)
        return exit_code

    def _print_completion_banner(self, exit_code: int, is_parallel: bool = False, details: dict = None):
        """Prints a highly professional execution banner."""
        print("\n" + "=" * 80)
        if exit_code == 0:
            print("[SUCCESS] ALL AUTOMATED TESTS PASSED SUCCESSFULLY!")
        else:
            print(f"[FAILURE] SOME TESTS FAILED (Exit Code: {exit_code})")
        print("=" * 80)
        
        if is_parallel and details:
            print("Parallel Scenario Run Details:")
            for tag, code in details.items():
                status = "PASSED" if code == 0 else "FAILED"
                print(f"  - Tag {tag:<10} : {status}")
            print("-" * 80)
            
        print("\nGenerated Outputs Map:")
        print(f"  - Excel test sheet   : {os.path.abspath('test_cases.xlsx')}")
        if is_parallel:
            print(f"  - Parallel HTMLs     : {os.path.abspath('reports/report_tc_pb_*.html')}")
            print(f"  - Parallel JSONs     : {os.path.abspath('reports/report_tc_pb_*.json')}")
        else:
            print(f"  - HTML Report        : {os.path.abspath('reports/report.html')}")
            print(f"  - JSON Report        : {os.path.abspath('reports/report.json')}")
        print(f"  - JUnit XML Directory: {os.path.abspath('reports/junit')}")
        print(f"  - Screenshots/Proof  : {os.path.abspath('proof')}")
        print("-" * 80 + "\n")

def main():
    # Set up argument parser for command line execution
    parser = argparse.ArgumentParser(description="ParaBank BDD Test Runner")
    parser.add_argument(
        "-t", "--tags",
        help="BDD tag expression to filter scenarios (e.g., '@positive', '@negative', '@registration')",
        default=None
    )
    parser.add_argument(
        "-p", "--parallel",
        help="Run scenarios in parallel processes using Playwright concurrency",
        action="store_true"
    )
    args, unknown = parser.parse_known_args()

    print("=" * 80)
    print("      PARABANK BDD TEST AUTOMATION SUITE: BROWSER & REPORTING RUNNER     ")
    print("=" * 80)
    print(f"[CONFIG] Project Root: {PROJECT_ROOT}")
    
    # Resolve tags: CLI Argument takes priority, followed by environment variables
    resolved_tags = args.tags or os.getenv("TAGS") or os.getenv("TEST_TAGS")
    
    # Resolve parallel run
    use_parallel = args.parallel or os.getenv("PARALLEL", "").lower() == "true"
    
    # Instantiate the runner class and execute the steps
    runner = ParaBankTestRunner()
    try:
        runner.generate_excel_documentation()
        runner.prepare_directories()
        
        if use_parallel:
            exit_code = runner.run_tests_in_parallel(resolved_tags)
        else:
            exit_code = runner.run_tests_by_tag(resolved_tags)
            
        sys.exit(exit_code)
    except Exception as e:
        print(f"[FATAL] Test Runner encountered a critical error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
